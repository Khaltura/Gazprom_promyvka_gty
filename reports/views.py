from collections import defaultdict
from decimal import Decimal
from io import BytesIO
import json

from django.contrib.auth.decorators import login_required
from django.db import models
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from balance.models import MaterialOperation
from balance.views import build_row
from promyvki.models import Antifreeze, CleaningFluid, CompressorStation, Wash
from users.utils import has_assigned_stations, user_stations

ZERO = Decimal("0.00")


def date_range(queryset, request):
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    return queryset


def scoped_stations(request):
    available = user_stations(request.user)
    stations = available
    selected = None
    station_id = request.GET.get("station")
    if station_id:
        selected = available.filter(pk=station_id).first()
        if selected:
            stations = available.filter(pk=selected.pk)
    return stations, selected


def scoped_washes(request, stations):
    return date_range(
        Wash.objects.filter(station__in=stations).select_related(
            "station", "cleaning_fluid", "antifreeze", "created_by"
        ),
        request,
    )


def scoped_operations(request, stations):
    return date_range(
        MaterialOperation.objects.filter(
            models.Q(from_station__in=stations) | models.Q(to_station__in=stations)
        )
        .select_related(
            "from_station", "to_station", "cleaning_fluid", "antifreeze", "created_by"
        )
        .distinct(),
        request,
    )


def access_or_pending(request):
    return has_assigned_stations(request.user)


@login_required
def analytics(request):
    return redirect("wash_analytics")


@login_required
def wash_analytics(request):
    if not access_or_pending(request):
        return redirect("account_pending")

    stations, selected = scoped_stations(request)
    washes = scoped_washes(request, stations)

    daily_usage = defaultdict(lambda: {"cleaning": 0.0, "antifreeze": 0.0})
    for wash in washes.order_by("date", "id"):
        key = wash.date.strftime("%d.%m.%Y")
        daily_usage[key]["cleaning"] += float(wash.cleaning_fluid_amount or 0)
        daily_usage[key]["antifreeze"] += float(wash.antifreeze_amount or 0)

    cleaning_by_type = list(
        washes.values("cleaning_fluid__name")
        .annotate(value=Sum("cleaning_fluid_amount"))
        .order_by("cleaning_fluid__name")
    )
    antifreeze_by_type = list(
        washes.exclude(antifreeze=None)
        .values("antifreeze__name")
        .annotate(value=Sum("antifreeze_amount"))
        .order_by("antifreeze__name")
    )

    materials_pie = [
        {
            "name": f"Промывочная жидкость · {row['cleaning_fluid__name']}",
            "value": float(row["value"] or 0),
        }
        for row in cleaning_by_type
        if (row["value"] or 0) > 0
    ]
    materials_pie.extend(
        {
            "name": f"Антифриз · {row['antifreeze__name']}",
            "value": float(row["value"] or 0),
        }
        for row in antifreeze_by_type
        if (row["value"] or 0) > 0
    )

    cards = []
    for station in stations:
        station_washes = washes.filter(station=station)
        cards.append(
            {
                "station": station,
                "washes": station_washes.count(),
                "cleaning": station_washes.aggregate(value=Sum("cleaning_fluid_amount"))["value"] or ZERO,
                "antifreeze": station_washes.aggregate(value=Sum("antifreeze_amount"))["value"] or ZERO,
                "last": station_washes.order_by("-date", "-id").first(),
            }
        )

    context = {
        "stations_all": user_stations(request.user),
        "selected_station": selected,
        "cards": cards,
        "wash_count": washes.count(),
        "cleaning_total": washes.aggregate(value=Sum("cleaning_fluid_amount"))["value"] or ZERO,
        "antifreeze_total": washes.aggregate(value=Sum("antifreeze_amount"))["value"] or ZERO,
        "daily_json": json.dumps(
            [{"date": key, **values} for key, values in daily_usage.items()],
            ensure_ascii=False,
        ),
        "materials_pie_json": json.dumps(materials_pie, ensure_ascii=False),
    }
    return render(request, "reports/wash_analytics.html", context)


@login_required
def balance_analytics(request):
    if not access_or_pending(request):
        return redirect("account_pending")

    stations, selected = scoped_stations(request)
    operations = scoped_operations(request, stations)

    daily = defaultdict(lambda: {"arrival": 0.0, "writeoff": 0.0, "transfer": 0.0})
    for operation in operations.order_by("date", "id"):
        key = operation.date.strftime("%d.%m.%Y")
        amount = float(operation.amount or 0)
        if operation.operation_type == MaterialOperation.ARRIVAL:
            daily[key]["arrival"] += amount
        elif operation.operation_type == MaterialOperation.WRITEOFF:
            daily[key]["writeoff"] += amount
        else:
            daily[key]["transfer"] += amount

    balances = []
    cleaning_balance = ZERO
    antifreeze_balance = ZERO
    for station in stations:
        for material_type, items in (
            (MaterialOperation.CLEANING, CleaningFluid.objects.all()),
            (MaterialOperation.ANTIFREEZE, Antifreeze.objects.all()),
        ):
            for item in items:
                row = build_row(station, material_type, item)
                if row["active"]:
                    balances.append(row)
                    if material_type == MaterialOperation.CLEANING:
                        cleaning_balance += row["balance"]
                    else:
                        antifreeze_balance += row["balance"]

    balance_chart = [
        {
            "name": f"{row['station']} · {row['material']}",
            "value": float(row["balance"]),
            "type": row["material_type_label"],
        }
        for row in balances
    ]

    context = {
        "stations_all": user_stations(request.user),
        "selected_station": selected,
        "operations_count": operations.count(),
        "arrival_total": operations.filter(operation_type=MaterialOperation.ARRIVAL).aggregate(value=Sum("amount"))["value"] or ZERO,
        "writeoff_total": operations.filter(operation_type=MaterialOperation.WRITEOFF).aggregate(value=Sum("amount"))["value"] or ZERO,
        "transfer_total": operations.filter(operation_type=MaterialOperation.TRANSFER).aggregate(value=Sum("amount"))["value"] or ZERO,
        "cleaning_balance": cleaning_balance,
        "antifreeze_balance": antifreeze_balance,
        "daily_json": json.dumps([{"date": key, **values} for key, values in daily.items()], ensure_ascii=False),
        "balance_json": json.dumps(balance_chart, ensure_ascii=False),
    }
    return render(request, "reports/balance_analytics.html", context)


def style_workbook(workbook):
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1565C0")
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = min(
                max(
                    14,
                    max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)) + 2,
                ),
                38,
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions


def add_balance_sheet(workbook, stations):
    sheet = workbook.active
    sheet.title = "Баланс"
    sheet.append(["КС", "Тип материала", "Материал", "Приход", "Получено", "Передано", "Списано", "Расход", "Остаток"])
    for station in stations:
        for material_type, items in (
            (MaterialOperation.CLEANING, CleaningFluid.objects.all()),
            (MaterialOperation.ANTIFREEZE, Antifreeze.objects.all()),
        ):
            for item in items:
                row = build_row(station, material_type, item)
                if row["active"]:
                    sheet.append([
                        station.name,
                        row["material_type_label"],
                        str(item),
                        float(row["arrival"]),
                        float(row["transfer_in"]),
                        float(row["transfer_out"]),
                        float(row["writeoff"]),
                        float(row["wash"]),
                        float(row["balance"]),
                    ])


def add_operations_sheet(workbook, operations):
    sheet = workbook.create_sheet("Операции")
    sheet.append(["Дата", "Тип", "Тип материала", "Материал", "Откуда", "Куда", "Количество", "Комментарий", "Пользователь"])
    for operation in operations:
        sheet.append([
            operation.date,
            operation.get_operation_type_display(),
            operation.get_material_type_display(),
            operation.material_name,
            str(operation.from_station or ""),
            str(operation.to_station or ""),
            float(operation.amount),
            operation.comment,
            str(operation.created_by or ""),
        ])


def add_washes_sheet(workbook, washes, use_active=False):
    sheet = workbook.active if use_active else workbook.create_sheet("Промывки")
    sheet.title = "Промывки"
    sheet.append(["Дата", "КС", "КТС до", "КТС после", "Жидкость", "Расход жидкости", "Антифриз", "Расход антифриза", "Пользователь"])
    for wash in washes:
        sheet.append([
            wash.date,
            wash.station.name,
            float(wash.kts_before),
            float(wash.kts_after),
            wash.cleaning_fluid.name,
            float(wash.cleaning_fluid_amount),
            str(wash.antifreeze or ""),
            float(wash.antifreeze_amount),
            str(wash.created_by or ""),
        ])


def excel_response(workbook, filename):
    style_workbook(workbook)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_excel(request):
    """Общий Excel из верхней панели: баланс, операции и промывки."""
    if not access_or_pending(request):
        return redirect("account_pending")
    stations, _ = scoped_stations(request)
    operations = scoped_operations(request, stations)
    washes = scoped_washes(request, stations)
    workbook = Workbook()
    add_balance_sheet(workbook, stations)
    add_operations_sheet(workbook, operations)
    add_washes_sheet(workbook, washes)
    return excel_response(workbook, "full_report.xlsx")


@login_required
def export_balance_excel(request):
    """Excel со страницы баланса: только баланс и операции."""
    if not access_or_pending(request):
        return redirect("account_pending")
    stations, _ = scoped_stations(request)
    operations = scoped_operations(request, stations)
    workbook = Workbook()
    add_balance_sheet(workbook, stations)
    add_operations_sheet(workbook, operations)
    return excel_response(workbook, "balance_and_operations.xlsx")


@login_required
def export_washes_excel(request):
    """Excel со страницы промывок: только промывки."""
    if not access_or_pending(request):
        return redirect("account_pending")
    stations, _ = scoped_stations(request)
    washes = scoped_washes(request, stations)
    workbook = Workbook()
    add_washes_sheet(workbook, washes, use_active=True)
    return excel_response(workbook, "washes.xlsx")
